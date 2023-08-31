import pcbnew
import os
from openpyxl import Workbook
# import wx
# import wx.html2

from pcbnew import *
import math

def distance(p1, p2):
    """Calculate the Euclidean distance between two points"""
    return math.sqrt((p1[0]-p2[0])**2 + (p1[1]-p2[1])**2)

def min_distance(points):
    """Find the minimum distance between two points in a list"""
    min_dist = float('inf')
    for i in range(len(points)):
        for j in range(i+1, len(points)):
            dist = distance(points[i], points[j])
            if dist < min_dist:
                min_dist = dist
    return min_dist

def get_pins(component):
    pinout = []
    added_pads = [] 
    for pad in component.Pads():
            if pad.GetNumber() not in added_pads and pad.GetNumber() != '': # filter redundant and bogus pads (eg.: thermal pad)
                pinout.append(pad)
                added_pads.append(pad.GetNumber())
    return pinout

def get_xy(pad):
    return pad.GetPosition()[0],pad.GetPosition()[1]

class BgaToExcel(pcbnew.ActionPlugin):
    def defaults(self):
        self.name = "BgaToExcel"
        self.category = "Documentation"
        self.description = "Outputs a BGA into excel in the current rotation"
        self.show_toolbar_button = True
        self.icon_file_name = os.path.join(os.path.dirname(__file__), '4-pin-BGA.png')



    def Run(self):
        
        #board = LoadBoard('C:/Users/samwi/Documents/Kicad_Projects/script_testing/script_testing.kicad_pcb')
        board = GetBoard()
        footprintSelected = 0
        outputString = ""

        # for footprint in pcbnew.GetBoard().GetFootprints():
        for footprint in board.GetFootprints():
            print(footprint.GetReference())

            if footprint.IsSelected():
            #if footprint.GetReference() == 'U2':
                print("footprint selected")
                footprintSelected = footprint
                #self.footprint_selection.append(footprint)
        

        pinout = get_pins(footprintSelected)

        sortedPads = sorted(pinout, key=get_xy)
        #get list of pad coordinates from sorted list of pads
        xy_list = [(get_xy(obj)[0], get_xy(obj)[1]) for obj in sortedPads]
        #convert to mm and round off to 3 decimal places
        xy_list_mm = [(round(t[0]/1000000.0,3), round(t[1]/1000000.0,3)) for t in xy_list]
        #calculate minimum distance between two pads this is finding the pitch of the part
        min_dist = round(min_distance(xy_list_mm),3)
        print("Minimum distance " + str(min_dist))
        #find the minimum and maximum x and y value to shift cells to 0,0
        min_x = min(coord[0] for coord in xy_list_mm)
        min_y = min(coord[1] for coord in xy_list_mm)
        print("Min x "+ str(min_x))
        print("Min y "+ str(min_y))


        # create a new Excel workbook
        workbook = Workbook()

        # select the active worksheet
        worksheet = workbook.active

        # Loop through the coordinates and write them to the sheet
        for pad in sortedPads:

            coord = get_xy(pad)
            print(coord)
            coord = (round(coord[0]/1000000.0,3), round(coord[1]/1000000.0,3))
            print(coord)
            coord = (round(coord[0] - min_x,2), round(coord[1] - min_y,2))
            print(coord)
            coord = (int(round(coord[0] / min_dist)), int(round(coord[1]/min_dist))) 
            print(coord)
            # Convert the x,y coordinates to Excel row,column coordinates
            row = coord[1] + 1
            col = coord[0] + 1

            input_str = pad.GetNet().GetNetname()

            # Get the indices of the second and third hyphens in the input string
            second_hyphen_index = input_str.find('-', input_str.find('-') + 1)
            third_hyphen_index = input_str.find('-', second_hyphen_index + 1)

            # Check if the second and third hyphens are found in the input string
            if second_hyphen_index != -1 and third_hyphen_index != -1:
                # Extract the substring between the second and third hyphens
                result = input_str[second_hyphen_index+1:third_hyphen_index]

            else:
                print("Error: Invalid input string.")
            # Write the coordinates to the sheet
            worksheet.cell(row=row, column=col, value=result)
            #worksheet.cell(row=row, column=col, value=pad.GetPadName())

        # # Loop through each cell in the first row and set the value
        # for col in worksheet.iter_cols(min_row=1, max_row=1):
        #     for cell in col:
        #         cell.value = "This is some text"

        # Loop through each column in the first row and set the column width
        for col in worksheet.iter_cols(min_row=1, max_row=1):
            for cell in col:
                worksheet.column_dimensions[cell.column_letter].auto_size = True

        # save the workbook to a file
        workbook.save(filename="C:/Users/samwi/Documents/Kicad_Projects/script_testing/data.xlsx")
        # wx.MessageBox(outputString, 'Measure of length of tracks', wx.OK | wx.ICON_INFORMATION)
        
BgaToExcel().register()

bga_to_excel = BgaToExcel()

bga_to_excel.Run()